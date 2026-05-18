"""
Inspect both sheets in the Excel file to understand structure
"""
import openpyxl

xlsx_path = r'E:\Sahifati\Public-data-\modify lines pages 1-100.xlsx'

wb = openpyxl.load_workbook(xlsx_path, data_only=True)
print(f"Sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data_rows = [r for r in rows[1:] if any(v is not None for v in r)]
    print(f"\n=== {sheet_name} ===")
    print(f"  Header: {header}")
    print(f"  Data rows: {len(data_rows)}")
    print(f"  First 5 rows:")
    for r in data_rows[:5]:
        print(f"    {r}")
    print(f"  Last 3 rows:")
    for r in data_rows[-3:]:
        print(f"    {r}")

wb.close()
