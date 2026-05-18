"""
Apply all corrections from both Excel sheets to mushaf_layout_mushaf5.json.
  Sheet1: word_index, line  -> update line number for these words
  Sheet2: word_index, page  -> move words to a different page

Source: original backup (pre-any-corrections)
"""
import openpyxl
import json
import datetime
import shutil
import os

xlsx_path  = r'E:\Sahifati\Public-data-\modify lines pages 1-100.xlsx'
json_src   = r'E:\Sahifati\frontend_users\ui\assets\json\mushaf_layout_mushaf5.json.bak_20260518_185339'
json_dest  = r'E:\Sahifati\frontend_users\ui\assets\json\mushaf_layout_mushaf5.json'

# ── 1. Read Excel ──────────────────────────────────────────────────────────
print("Reading Excel corrections...")
wb = openpyxl.load_workbook(xlsx_path, data_only=True)

sheet1_rows = list(wb['Sheet1'].iter_rows(values_only=True))
sheet1_corrections = {}   # word_index -> new_line
for row in sheet1_rows[1:]:
    if row[0] is None: continue
    try:   sheet1_corrections[int(row[0])] = int(row[1])
    except (TypeError, ValueError): pass

sheet2_rows = list(wb['Sheet2'].iter_rows(values_only=True))
sheet2_corrections = {}   # word_index -> new_page_number (1-based)
for row in sheet2_rows[1:]:
    if row[0] is None: continue
    try:   sheet2_corrections[int(row[0])] = int(row[1])
    except (TypeError, ValueError): pass

wb.close()
print(f"  Sheet1 (line fixes):    {len(sheet1_corrections)} entries")
print(f"  Sheet2 (page moves):    {len(sheet2_corrections)} entries")

overlap = set(sheet1_corrections) & set(sheet2_corrections)
print(f"  Words in both sheets:   {len(overlap)} (will use Sheet1 line for these)")

# ── 2. Load original JSON and flatten ─────────────────────────────────────
print(f"\nReading source JSON: {os.path.basename(json_src)}")
with open(json_src, 'r', encoding='utf-8') as f:
    data = json.load(f)

flat_words = []   # {wi, page_idx, page_num, line, word}
wi = 0
for page_idx, page in enumerate(data['pages']):
    pg_num = page['pageNumber']
    for line in page['lines']:
        ln = line['lineNumber']
        for word in line['words']:
            wi += 1
            flat_words.append({'wi': wi, 'page_idx': page_idx,
                                'page_num': pg_num, 'line': ln, 'word': word})

print(f"  Words indexed: {wi}")

# ── 3. Apply Sheet1 line corrections ──────────────────────────────────────
s1_changed = 0
for fw in flat_words:
    if fw['wi'] in sheet1_corrections:
        new_ln = sheet1_corrections[fw['wi']]
        if fw['line'] != new_ln:
            fw['line'] = new_ln
            s1_changed += 1
print(f"  Sheet1 changes applied: {s1_changed}")

# ── 4. Apply Sheet2 page reassignments ────────────────────────────────────
# Build page_num -> page_idx lookup
page_num_to_idx = {page['pageNumber']: idx for idx, page in enumerate(data['pages'])}

s2_changed = 0
s2_warnings = []
for fw in flat_words:
    if fw['wi'] in sheet2_corrections:
        new_pg_num = sheet2_corrections[fw['wi']]
        if new_pg_num not in page_num_to_idx:
            s2_warnings.append(f"  WARNING: word {fw['wi']} -> page {new_pg_num} not found in JSON!")
            continue
        new_pg_idx = page_num_to_idx[new_pg_num]
        if fw['page_idx'] != new_pg_idx:
            old_pg = fw['page_num']
            fw['page_idx'] = new_pg_idx
            fw['page_num'] = new_pg_num
            s2_changed += 1
            print(f"  Move wi={fw['wi']:5d}: page {old_pg} -> {new_pg_num}  line={fw['line']}")

for w in s2_warnings:
    print(w)
print(f"  Sheet2 page moves applied: {s2_changed}")

# ── 5. Rebuild JSON ────────────────────────────────────────────────────────
print("\nRebuilding JSON structure...")

for page_idx, page in enumerate(data['pages']):
    page_words = [fw for fw in flat_words if fw['page_idx'] == page_idx]
    line_groups = {}
    for fw in page_words:
        ln = fw['line']
        if ln not in line_groups:
            line_groups[ln] = []
        line_groups[ln].append(fw)
    for ln in line_groups:
        line_groups[ln].sort(key=lambda x: x['wi'])
    new_lines = [{'lineNumber': ln, 'words': [fw['word'] for fw in line_groups[ln]]}
                 for ln in sorted(line_groups.keys())]
    page['lines'] = new_lines

# ── 6. Verify total word count ─────────────────────────────────────────────
total = sum(len(l['words']) for p in data['pages'] for l in p['lines'])
print(f"Total words after rebuild: {total} (expected {wi})")
if total != wi:
    print("ERROR: word count mismatch! Aborting save.")
    exit(1)

# ── 7. Backup current JSON and save ───────────────────────────────────────
ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
if os.path.exists(json_dest):
    backup_path = json_dest + f'.bak_{ts}'
    shutil.copy2(json_dest, backup_path)
    print(f"Backup saved: {os.path.basename(backup_path)}")

with open(json_dest, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, separators=(',', ':'))

file_size = os.path.getsize(json_dest)
print(f"JSON saved: {file_size:,} bytes")
print(f"\nDone! Sheet1={s1_changed} line fixes, Sheet2={s2_changed} page moves.")
